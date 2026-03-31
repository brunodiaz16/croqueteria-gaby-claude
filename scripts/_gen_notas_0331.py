import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

FECHA = "2026-03-31"
OUTPUT_DIR = "data/xlsx"

pedidos = {
    "Dartacan": [
        ("Cat Chow Mariscos 20kg",       7),
        ("Campeon Recetas Caseras 25kg", 1),
        ("Gatina 15kg",                  1),
    ],
    "Chapetes": [
        ("Chapetes Premium 18kg",        1),
    ],
    "Costco": [
        ("Kirkland Salmon/Camote 15.87kg", 1),
    ],
}

COLORS = {
    "Chapetes": "1A7A6E",
    "Dartacan": "1F4E79",
    "Invet":    "5B2C8D",
    "Costco":   "C05700",
}

def b():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_sheet(ws, proveedor, productos):
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    color = COLORS.get(proveedor, "333333")
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    fecha_display = datetime.strptime(FECHA, "%Y-%m-%d").strftime("%d %b %Y").upper()

    ws.merge_cells('A1:C1')
    ws['A1'] = f"PEDIDO {proveedor.upper()} — {fecha_display}"
    ws['A1'].font = Font(bold=True, size=13, color=color)
    ws['A1'].alignment = Alignment(horizontal="center")

    for col, label in enumerate(['Producto', 'Cant. Vendida', 'Cant. a Pedir'], 1):
        c = ws.cell(row=2, column=col, value=label)
        c.fill = fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = b()

    for row, (prod, qty) in enumerate(productos, start=3):
        ws.cell(row=row, column=1, value=prod).border = b()
        for c in [2, 3]:
            cell = ws.cell(row=row, column=c, value=qty)
            cell.border = b()
            cell.alignment = Alignment(horizontal="center")
        if qty >= 2:
            for c in range(1, 4):
                ws.cell(row=row, column=c).font = Font(bold=True)

# Individuales
for prov, prods in pedidos.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido"
    make_sheet(ws, prov, prods)
    wb.save(f"{OUTPUT_DIR}/Nota_Pedido_{prov}_{FECHA}.xlsx")
    print(f"Generado: Nota_Pedido_{prov}_{FECHA}.xlsx")

# General
wb_gen = openpyxl.Workbook()
wb_gen.remove(wb_gen.active)
for prov, prods in pedidos.items():
    ws = wb_gen.create_sheet(title=prov)
    make_sheet(ws, prov, prods)
wb_gen.save(f"{OUTPUT_DIR}/Nota_Pedido_General_{FECHA}.xlsx")
print(f"Generado: Nota_Pedido_General_{FECHA}.xlsx")
