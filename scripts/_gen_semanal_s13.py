import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = "data/xlsx/Reporte_Semanal_S13_2026.xlsx"
AZUL  = "1F4E79"
VERDE = "1A7A6E"
GRIS  = "F2F2F2"

dias = [
    ("Lun 23 Mar", 23, 16643.0,  1220.0, 7.3,  1),
    ("Mar 24 Mar", 10,  7906.0,   925.0, 11.7, 0),
    ("Mié 25 Mar", 16, 13673.0,  1592.0, 11.6, 1),
    ("Jue 26 Mar", 22, 17948.0,  1644.0, 9.2,  0),
    ("Vie 27 Mar", 13, 12469.0,  1380.0, 11.1, 3),
]
total_ord  = sum(d[1] for d in dias)
total_neto = 68638.81
total_gan  = 6760.26
total_flex = sum(d[5] for d in dias)

gastos_lista = [
    ("gasolina",          "Gasolina semana",              1200.0),
    ("comidas",           "Comida trabajador",              390.0),
    ("pago_trabajador",   "Trabajador (base+5 Flex)",      2950.0),
    ("compra_inventario", "LiveClear x10 (*inv. stock)",   3700.0),
]
total_gastos_op  = 1200 + 390 + 2950
total_gastos_all = 8240.0
gan_real    = total_gan - total_gastos_op
margen_real = round(gan_real / total_neto * 100, 1)

top_prods = [
    ("Cat Chow Mariscos 20kg",         7,  6596, 6.1,  "ROJO"),
    ("Arena Scoop Away 19kg 2-Pack",   7,  6272, 10.9, "AMARILLO"),
    ("Dog Chow 25kg Raza Pequena",     5,  5016, 10.3, "AMARILLO"),
    ("Campeon Adulto 25kg",            6,  4526, 0.6,  "ROJO"),
    ("Ganador Premium 20kg v2",        4,  4186, 7.3,  "ROJO"),
    ("Perron Adulto 25kg 2-Pack",      3,  3490, 8.0,  "AMARILLO"),
    ("Ganador Premium 21kg",           3,  3163, 8.0,  "AMARILLO"),
    ("Perron Adulto 25kg",             4,  2825, 24.2, "VERDE"),
    ("Gatina 15kg",                    5,  2733, 9.4,  "AMARILLO"),
    ("Chapetes Premium 18kg",          4,  2005, 18.2, "VERDE"),
]
alertas = [
    ("Dog Chow Croquetas 25kg",    -1.0,  891, "PAUSAR - venta a perdida"),
    ("Campeon Adulto 25kg",         0.6, 4526, "REPRECIAR URGENTE - 6u en rojo"),
    ("Kan Kan 25kg",                3.2,  393, "Margen critico"),
    ("Perron Premium Adulto",       4.0,  724, "Margen critico"),
    ("Maskottchen Premium 15kg",    4.2,  548, "Margen bajo"),
    ("Cat Chow Mariscos 20kg",      6.1, 6596, "Estrella pero en rojo - repreciar"),
    ("Ganador Premium 20kg v2",     7.3, 4186, "Repreciar urgente"),
    ("Gatina 15kg",                 9.4, 2733, "Monitorear - borderline"),
]
compras_sug = [
    ("Cat Chow Mariscos 20kg",        "1.4/dia", "Dartacan",  "Alta rotacion - repreciar antes de restock"),
    ("Arena Scoop Away 19kg 2-Pack",  "1.4/dia", "Costco",    "Alta rotacion - mantener stock"),
    ("Dog Chow 25kg Raza Pequena",    "1.0/dia", "Dartacan",  "Nuevo listing - monitorear"),
    ("Campeon Adulto 25kg",           "1.2/dia", "Dartacan",  "PAUSAR o repreciar antes de restock"),
    ("Ganador Premium 20kg",          "0.8/dia", "Dartacan",  "Repreciar antes de restock"),
    ("Perron Adulto 25kg 2-Pack",     "0.6/dia", "Dartacan",  "Margen borderline OK"),
    ("Gatina 15kg",                   "1.0/dia", "Dartacan",  "Monitorear envios divididos ML"),
    ("Chapetes Premium 18kg",         "0.8/dia", "Chapetes",  "Buen margen - mantener stock"),
]

def b():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg=AZUL):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = b()
    return c

def cell(ws, row, col, val, bold=False, align="left", bg=None, color=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=color or "000000")
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = b()
    return c

wb = openpyxl.Workbook()

# RESUMEN
ws1 = wb.active
ws1.title = "Resumen"
ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 30

ws1.merge_cells("A1:C1")
ws1["A1"] = "REPORTE SEMANAL S13 2026 — 23 al 27 Marzo"
ws1["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws1["A1"].fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 26

rows_res = [
    ("VENTAS",          "",                                       True,  GRIS),
    ("Ordenes totales", str(total_ord),                           False, None),
    ("Neto ML semana",  f"${total_neto:,.2f}",                   False, None),
    ("Ganancia bruta",  f"${total_gan:,.2f}",                    False, None),
    ("Margen bruto",    f"{round(total_gan/total_neto*100,1)}%", False, None),
    ("Envios Flex",     str(total_flex),                          False, None),
    ("GASTOS",          "",                                       True,  GRIS),
    ("Gasolina",        "$1,200.00",                              False, None),
    ("Comidas",         "$390.00",                                False, None),
    ("Pago trabajador", "$2,950.00",                              False, None),
    ("LiveClear stock*","$3,700.00",                              False, None),
    ("Total gastos op.","$4,540.00",                              True,  None),
    ("(*sin inv. stock)","",                                      False, None),
    ("RESULTADO",       "",                                       True,  GRIS),
    ("Ganancia real",   f"${gan_real:,.2f}",                     True,  None),
    ("Margen real",     f"{margen_real}%",                        True,  None),
    ("*LiveClear es inversion de inventario, no gasto recurrente", "", False, None),
]
for i, (lbl, val, bold, bg) in enumerate(rows_res, start=2):
    cell(ws1, i, 1, lbl, bold=bold, bg=bg or "FFFFFF")
    c = cell(ws1, i, 2, val, bold=bold, bg=bg or "FFFFFF", align="right")
    if "Ganancia real" in lbl or "Margen real" in lbl:
        c.font = Font(bold=True, color="00A651" if gan_real >= 0 else "C00000")

# POR DIA
ws2 = wb.create_sheet("Por_Dia")
for col, w in zip("ABCDEFG", [14,10,14,14,12,8,16]):
    ws2.column_dimensions[col].width = w
for ci, h in enumerate(["Dia","Ordenes","Neto ML","Ganancia","Margen","Flex","Neto Acum."], 1):
    hdr(ws2, 1, ci, h)
acum = 0
for ri, (dia, ord_, neto, gan, margen, flex) in enumerate(dias, start=2):
    acum += neto
    sem = "VERDE" if margen > 14 else ("AMARILLO" if margen >= 8 else "ROJO")
    bg = "E2EFDA" if sem=="VERDE" else ("FFF2CC" if sem=="AMARILLO" else "FFE0E0")
    for ci, v in enumerate([dia, ord_, f"${neto:,.0f}", f"${gan:,.0f}", f"{margen}%", flex, f"${acum:,.0f}"], 1):
        cell(ws2, ri, ci, v, align="center", bg=bg)
for ci, v in enumerate(["TOTAL", total_ord, f"${total_neto:,.0f}", f"${total_gan:,.0f}",
                         f"{round(total_gan/total_neto*100,1)}%", total_flex, ""], 1):
    cell(ws2, len(dias)+2, ci, v, bold=True, align="center", bg=GRIS)

# TOP PRODUCTOS
ws3 = wb.create_sheet("Top_Productos")
ws3.column_dimensions["A"].width = 36
for col, w in zip("BCDE", [10,14,12,12]):
    ws3.column_dimensions[col].width = w
for ci, h in enumerate(["Producto","Unidades","Neto Total","Margen","Semaforo"], 1):
    hdr(ws3, 1, ci, h, bg=VERDE)
for ri, (prod, u, neto, margen, sem) in enumerate(top_prods, start=2):
    bg = "E2EFDA" if sem=="VERDE" else ("FFF2CC" if sem=="AMARILLO" else "FFE0E0")
    cell(ws3, ri, 1, prod, bg=bg)
    cell(ws3, ri, 2, u, align="center", bg=bg)
    cell(ws3, ri, 3, f"${neto:,}", align="right", bg=bg)
    cell(ws3, ri, 4, f"{margen}%", align="center", bg=bg)
    cell(ws3, ri, 5, sem, align="center", bg=bg)

# ALERTAS
ws4 = wb.create_sheet("Alertas")
ws4.column_dimensions["A"].width = 34
for col, w in zip("BCD", [12,14,32]):
    ws4.column_dimensions[col].width = w
for ci, h in enumerate(["Producto","Margen","Neto Semana","Accion"], 1):
    hdr(ws4, 1, ci, h, bg="C00000")
for ri, (prod, margen, neto, accion) in enumerate(alertas, start=2):
    bg = "FFE0E0" if margen < 0 else ("FFE0E0" if margen < 8 else "FFF2CC")
    bold = margen < 0 or (margen < 1)
    cell(ws4, ri, 1, prod, bold=bold, bg=bg)
    cell(ws4, ri, 2, f"{margen}%", align="center", bg=bg, bold=bold)
    cell(ws4, ri, 3, f"${neto:,}", align="right", bg=bg)
    cell(ws4, ri, 4, accion, bg=bg)

# GASTOS
ws5 = wb.create_sheet("Gastos")
ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 32
ws5.column_dimensions["C"].width = 14
for ci, h in enumerate(["Categoria","Descripcion","Monto"], 1):
    hdr(ws5, 1, ci, h)
for ri, (cat, desc, monto) in enumerate(gastos_lista, start=2):
    cell(ws5, ri, 1, cat)
    cell(ws5, ri, 2, desc)
    cell(ws5, ri, 3, f"${monto:,.2f}", align="right")
cell(ws5, len(gastos_lista)+2, 2, "TOTAL", bold=True)
cell(ws5, len(gastos_lista)+2, 3, f"${total_gastos_all:,.2f}", bold=True, align="right")
cell(ws5, len(gastos_lista)+3, 2, "TOTAL (sin inventario)", bold=True)
cell(ws5, len(gastos_lista)+3, 3, f"${total_gastos_op:,.2f}", bold=True, align="right")

# COMPRAS SUGERIDAS
ws6 = wb.create_sheet("Compras_Sugeridas")
ws6.column_dimensions["A"].width = 34
for col, w in zip("BCD", [12,14,36]):
    ws6.column_dimensions[col].width = w
for ci, h in enumerate(["Producto","Vel. Venta","Proveedor","Nota"], 1):
    hdr(ws6, 1, ci, h, bg=VERDE)
for ri, (prod, vel, prov, nota) in enumerate(compras_sug, start=2):
    cell(ws6, ri, 1, prod)
    cell(ws6, ri, 2, vel, align="center")
    cell(ws6, ri, 3, prov, align="center")
    cell(ws6, ri, 4, nota)

wb.save(OUTPUT)
print(f"Generado: {OUTPUT}")
