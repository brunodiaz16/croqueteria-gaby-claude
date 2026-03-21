"""
Módulo utilitario para leer/escribir el Catálogo Maestro en Google Sheets.
Fallback a context/costos.md si el Sheet no está disponible.

Uso:
  from scripts.catalogo import leer_catalogo, leer_aliases, actualizar_costo
"""

import sys
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from scripts.config import (
    CATALOGO_SHEET_ID, SCOPES,
    CREDENTIALS_FILE, TOKEN_FILE, COSTOS_PATH, XLSX_DIR,
)


def _get_credentials():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                print(f"ERROR: No se encontro {CREDENTIALS_FILE}")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_FILE), SCOPES
            )
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json())
    return creds


def _get_sheets_service():
    creds = _get_credentials()
    return build("sheets", "v4", credentials=creds)


def leer_catalogo(sheet_id=None):
    """Lee la hoja Productos y retorna dict por publicacion_ml.

    Returns:
        dict: {publicacion_ml: {producto, marca, peso_kg, proveedor,
               costo_actual, categoria, ...}}
    """
    sid = sheet_id or CATALOGO_SHEET_ID
    if not sid:
        print("WARN: CATALOGO_SHEET_ID no configurado, usando fallback markdown")
        return _fallback_costos()

    try:
        service = _get_sheets_service()
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Productos!A:K")
            .execute()
        )
        rows = result.get("values", [])
        if len(rows) < 2:
            return {}

        headers = rows[0]
        catalogo = {}
        for row in rows[1:]:
            entry = {}
            for i, h in enumerate(headers):
                entry[h] = row[i] if i < len(row) else ""
            # Convertir costo a float
            try:
                entry["costo_actual"] = float(entry.get("costo_actual", 0))
            except (ValueError, TypeError):
                entry["costo_actual"] = None
            # Indexar por cada publicacion ML (pueden ser varias separadas por coma)
            pubs = str(entry.get("publicacion_ml", "")).split(",")
            for pub in pubs:
                pub = pub.strip()
                if pub:
                    catalogo[pub] = entry
        return catalogo
    except Exception as e:
        print(f"WARN: No se pudo leer Sheet: {e}. Usando fallback markdown.")
        return _fallback_costos()


def leer_aliases(sheet_id=None):
    """Lee la hoja Aliases y retorna dict {alias: {producto, proveedor}}.
    """
    sid = sheet_id or CATALOGO_SHEET_ID
    if not sid:
        return {}

    try:
        service = _get_sheets_service()
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Aliases!A:C")
            .execute()
        )
        rows = result.get("values", [])
        if len(rows) < 2:
            return {}

        aliases = {}
        for row in rows[1:]:
            if len(row) >= 2:
                aliases[row[0]] = {
                    "producto": row[1],
                    "proveedor": row[2] if len(row) > 2 else "",
                }
        return aliases
    except Exception as e:
        print(f"WARN: No se pudo leer aliases: {e}")
        return {}


def actualizar_costo(producto, nuevo_costo, proveedor, fuente, sheet_id=None):
    """Actualiza costo en hoja Productos y agrega entrada al Historial."""
    sid = sheet_id or CATALOGO_SHEET_ID
    if not sid:
        print("ERROR: CATALOGO_SHEET_ID no configurado")
        return False

    try:
        service = _get_sheets_service()

        # Leer productos para encontrar la fila
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Productos!A:K")
            .execute()
        )
        rows = result.get("values", [])
        headers = rows[0] if rows else []

        col_producto = headers.index("producto") if "producto" in headers else 0
        col_costo = headers.index("costo_actual") if "costo_actual" in headers else 4
        col_anterior = (
            headers.index("costo_anterior") if "costo_anterior" in headers else 5
        )
        col_fecha = headers.index("fecha_costo") if "fecha_costo" in headers else 6

        from datetime import date

        hoy = date.today().isoformat()

        for i, row in enumerate(rows[1:], start=2):
            nombre = row[col_producto] if col_producto < len(row) else ""
            if nombre.lower() == producto.lower():
                costo_ant = row[col_costo] if col_costo < len(row) else ""
                # Actualizar costo_actual, costo_anterior, fecha (en orden de columna)
                col_start = min(col_costo, col_anterior)
                col_end = col_fecha
                # Construir fila completa en orden: costo_actual=E, costo_anterior=F, fecha=G
                values_map = {col_costo: nuevo_costo, col_anterior: costo_ant, col_fecha: hoy}
                vals = [values_map[c] for c in range(col_start, col_end + 1)]
                service.spreadsheets().values().update(
                    spreadsheetId=sid,
                    range=f"Productos!{chr(65+col_start)}{i}:{chr(65+col_end)}{i}",
                    valueInputOption="RAW",
                    body={"values": [vals]},
                ).execute()

                # Agregar al historial
                service.spreadsheets().values().append(
                    spreadsheetId=sid,
                    range="Historial_Costos!A:F",
                    valueInputOption="RAW",
                    body={
                        "values": [
                            [hoy, producto, costo_ant, nuevo_costo, proveedor, fuente]
                        ]
                    },
                ).execute()
                print(f"  Catalogo actualizado: {producto} -> ${nuevo_costo}")
                return True

        print(f"  WARN: Producto '{producto}' no encontrado en el Sheet")
        return False
    except Exception as e:
        print(f"ERROR actualizando catalogo: {e}")
        return False


def _fallback_costos():
    """Lee costos desde context/costos.md como fallback."""
    costos_file = COSTOS_PATH
    if not costos_file.exists():
        return {}

    catalogo = {}
    lines = costos_file.read_text(encoding="utf-8").splitlines()
    in_table = False
    for line in lines:
        if line.startswith("| Producto"):
            in_table = True
            continue
        if in_table and line.startswith("|---"):
            continue
        if in_table and line.startswith("|"):
            parts = [p.strip() for p in line.split("|")[1:-1]]
            if len(parts) >= 4:
                producto = parts[0]
                costo_str = parts[1].replace("$", "").replace(",", "").strip()
                proveedor = parts[2]
                try:
                    costo = float(costo_str)
                except ValueError:
                    costo = None
                catalogo[producto] = {
                    "producto": producto,
                    "costo_actual": costo,
                    "proveedor": proveedor,
                }
        elif in_table and not line.startswith("|"):
            in_table = False

    return catalogo


def exportar_costos_md(sheet_id=None):
    """Exporta el catalogo del Sheet a context/costos.md como backup."""
    sid = sheet_id or CATALOGO_SHEET_ID
    if not sid:
        print("ERROR: CATALOGO_SHEET_ID no configurado")
        return False

    try:
        service = _get_sheets_service()

        # Leer productos
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Productos!A:K")
            .execute()
        )
        rows = result.get("values", [])

        # Leer aliases
        alias_result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Aliases!A:C")
            .execute()
        )
        alias_rows = alias_result.get("values", [])

        # Leer historial
        hist_result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sid, range="Historial_Costos!A:F")
            .execute()
        )
        hist_rows = hist_result.get("values", [])

        # Generar markdown
        lines = ["# COSTOS VIGENTES (auto-generado desde Google Sheet)"]
        lines.append("# NO EDITAR MANUALMENTE — usar scripts/catalogo.py")
        lines.append("")
        lines.append("| Producto | Costo | Proveedor | Actualizado |")
        lines.append("|---|---|---|---|")
        for row in rows[1:]:
            nombre = row[0] if len(row) > 0 else ""
            costo = row[4] if len(row) > 4 else ""
            prov = row[3] if len(row) > 3 else ""
            fecha = row[6] if len(row) > 6 else ""
            costo_str = f"${costo}" if costo else "SIN COSTO"
            lines.append(f"| {nombre} | {costo_str} | {prov} | {fecha} |")

        lines.append("")
        lines.append("# ALIASES")
        lines.append("| Ticket dice | Producto real | Proveedor |")
        lines.append("|---|---|---|")
        for row in alias_rows[1:]:
            alias = row[0] if len(row) > 0 else ""
            prod = row[1] if len(row) > 1 else ""
            prov = row[2] if len(row) > 2 else ""
            lines.append(f"| {alias} | {prod} | {prov} |")

        lines.append("")
        lines.append("# HISTORIAL DE CAMBIOS")
        lines.append("| Fecha | Producto | Antes | Ahora | Proveedor | Fuente |")
        lines.append("|---|---|---|---|---|---|")
        for row in hist_rows[1:]:
            while len(row) < 6:
                row.append("")
            lines.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} |")

        lines.append("")

        COSTOS_PATH.write_text("\n".join(lines), encoding="utf-8")
        print(f"Exportado: {COSTOS_PATH} ({len(rows)-1} productos)")
        return True
    except Exception as e:
        print(f"ERROR exportando costos: {e}")
        return False


def exportar_catalogo_xlsx(sheet_id=None, fecha=None):
    """Exporta el Catalogo_Maestro como XLSX a data/xlsx/ para subir a Drive.
    Genera: data/xlsx/Catalogo_Maestro_YYYY-MM-DD.xlsx
    """
    sid = sheet_id or CATALOGO_SHEET_ID
    hoy = fecha or date.today().strftime("%Y-%m-%d")
    if not sid:
        print("ERROR: CATALOGO_SHEET_ID no configurado")
        return None

    try:
        service = _get_sheets_service()

        def leer_hoja(rango):
            r = service.spreadsheets().values().get(spreadsheetId=sid, range=rango).execute()
            return r.get("values", [])

        productos = leer_hoja("Productos!A:K")
        aliases = leer_hoja("Aliases!A:C")
        historial = leer_hoja("Historial_Costos!A:F")

        wb = openpyxl.Workbook()
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(bold=True, color="FFFFFF")

        def escribir_hoja(ws, rows):
            if not rows:
                return
            for ci, val in enumerate(rows[0], 1):
                cell = ws.cell(row=1, column=ci, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(rows[1:], 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)

        ws_prod = wb.active
        ws_prod.title = "Productos"
        escribir_hoja(ws_prod, productos)

        ws_alias = wb.create_sheet("Aliases")
        escribir_hoja(ws_alias, aliases)

        ws_hist = wb.create_sheet("Historial_Costos")
        escribir_hoja(ws_hist, historial)

        XLSX_DIR.mkdir(parents=True, exist_ok=True)
        out = XLSX_DIR / f"Catalogo_Maestro_{hoy}.xlsx"
        wb.save(str(out))
        print(f"Catalogo exportado: {out}")
        return str(out)
    except Exception as e:
        print(f"ERROR exportando catalogo xlsx: {e}")
        return None


if __name__ == "__main__":
    print("Leyendo catalogo...")
    cat = leer_catalogo()
    if cat:
        for k, v in cat.items():
            print(f"  {k}: {v.get('producto')} -> ${v.get('costo_actual')}")
    else:
        print("  Catalogo vacio o no disponible")

    print("\nExportando costos.md...")
    exportar_costos_md()
