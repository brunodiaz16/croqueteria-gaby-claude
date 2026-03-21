"""Consolida reportes S12 (17-20 Mar 2026). Usa costos del reporte; solo busca en Sheet para SIN COSTO."""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

from scripts.catalogo import leer_catalogo

cat = leer_catalogo()
sheet_costos = {}
for v in cat.values():
    nombre = v.get("producto", "").lower().strip()
    costo  = v.get("costo_actual")
    if nombre and costo:
        sheet_costos[nombre] = float(costo)


def buscar_costo_sheet(titulo):
    """Solo para SIN COSTO: busca por coincidencia de palabras clave."""
    titulo_l = titulo.lower()
    if titulo_l in sheet_costos:
        return sheet_costos[titulo_l]
    best, best_score = None, 0
    for nombre_sheet, costo in sheet_costos.items():
        palabras = [p for p in nombre_sheet.split() if len(p) > 4]
        if not palabras:
            continue
        hits  = sum(1 for p in palabras if p in titulo_l)
        score = hits / len(palabras)
        if score > best_score and score >= 0.6:
            best_score = score
            best = costo
    return best


def multiplicador(titulo):
    t = titulo.lower()
    if "6 latas proplan gastro" in t:
        return 1
    if any(p in t for p in ["2 costales", "2 bultos", "2 bolsas", "2 pack", "2-pack", "paquete 2", "x2 costales"]):
        return 2
    if any(p in t for p in ["3 costales", "3 bultos"]):
        return 3
    if "24 sobres" in t:
        return 24
    return 1


def leer_reporte(fecha):
    path = Path(f"data/xlsx/Reporte_CroqueteriaGaby_{fecha}.xlsx")
    wb = openpyxl.load_workbook(str(path), data_only=True)

    if "Ventas" in wb.sheetnames:
        ws = wb["Ventas"]
        # leer headers para identificar columnas
        headers = [str(ws.cell(1, c).value or "").lower() for c in range(1, ws.max_column + 1)]
        # columnas esperadas: Producto, Pub ML, Unidades, Neto MXN / Neto_ML, Costo, Ganancia, Margen, Semaforo
        def col(keyword):
            for i, h in enumerate(headers):
                if keyword in h:
                    return i
            return None

        c_prod  = col("producto") or 0
        c_units = col("unidad")   or 2
        c_neto  = col("neto")     or 3
        c_costo = col("costo")    or 4
        c_sem   = col("semaf")    or 7

        filas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            prod = row[c_prod]
            if not prod or "TOTAL" in str(prod).upper():
                continue
            units = row[c_units] or 1
            neto  = row[c_neto]  or 0
            costo_rep = row[c_costo]
            sem   = str(row[c_sem] or "")
            if not neto:
                continue

            if sem == "SIN COSTO" or not costo_rep:
                # intentar desde Sheet
                c_base = buscar_costo_sheet(str(prod))
                mult   = multiplicador(str(prod))
                costo  = (c_base * mult) if c_base else None
            else:
                costo = float(costo_rep)

            filas.append({"prod": str(prod), "units": int(units), "neto": float(neto), "costo": costo, "sem": sem})

    else:
        # formato antiguo (Órdenes): no tiene costo — leer del Sheet
        for sn in wb.sheetnames:
            if "rden" in sn or "rden" in sn.lower():
                ws = wb[sn]
                break
        filas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            prod = row[2]
            if not prod:
                continue
            units = row[4] or 1
            neto  = row[8] or 0
            if not neto:
                continue
            c_base = buscar_costo_sheet(str(prod))
            mult   = multiplicador(str(prod))
            costo  = (c_base * mult) if c_base else None
            filas.append({"prod": str(prod), "units": int(units), "neto": float(neto), "costo": costo, "sem": ""})

    return filas


# ── CONSOLIDAR ──────────────────────────────────────────────────────────────────
DIAS = ["2026-03-17", "2026-03-18", "2026-03-19", "2026-03-20"]

resumen_dia = []
agg = defaultdict(lambda: {"u": 0, "neto": 0.0, "costo": 0.0})
sin_costo_global = []

for fecha in DIAS:
    filas = leer_reporte(fecha)
    neto_dia = gan_dia = 0.0
    sc_dia   = []
    for f in filas:
        neto_dia += f["neto"]
        if f["costo"]:
            gan_dia += f["neto"] - f["costo"]
        else:
            sc_dia.append(f["prod"][:40])
        agg[f["prod"]]["u"]    += f["units"]
        agg[f["prod"]]["neto"] += f["neto"]
        agg[f["prod"]]["costo"] += f["costo"] or 0

    marg = gan_dia / neto_dia * 100 if neto_dia else 0
    resumen_dia.append((fecha, len(filas), neto_dia, gan_dia, marg, sc_dia))
    sin_costo_global.extend(sc_dia)

    tag = f" | SIN COSTO: {sc_dia}" if sc_dia else ""
    print(f"{fecha}: {len(filas)} ords | neto ${neto_dia:,.2f} | gan ${gan_dia:,.2f} | {marg:.1f}%{tag}")

total_ords = sum(r[1] for r in resumen_dia)
total_neto = sum(r[2] for r in resumen_dia)
total_gan  = sum(r[3] for r in resumen_dia)
margen_p   = total_gan / total_neto * 100

print(f"\nSEMANA: {total_ords} ords | ${total_neto:,.2f} neto | ${total_gan:,.2f} gan | {margen_p:.1f}% margen")
if sin_costo_global:
    print(f"\nSIN COSTO pendientes: {list(set(sin_costo_global))}")

print("\n=== TOP por NETO ===")
for prod, d in sorted(agg.items(), key=lambda x: x[1]["neto"], reverse=True)[:12]:
    if d["costo"]:
        m    = (d["neto"] - d["costo"]) / d["neto"] * 100
        flag = "V" if m >= 14 else ("A" if m >= 8 else "R")
        print(f"  {prod[:48]:<48} x{d['u']:>2} ${d['neto']:>9,.0f}  {m:>5.1f}% [{flag}]")
    else:
        print(f"  {prod[:48]:<48} x{d['u']:>2} ${d['neto']:>9,.0f}  SIN COSTO")

# guardar JSON para generar_semanal.py
Path("data").mkdir(exist_ok=True)
Path("data/semanal_data.json").write_text(
    json.dumps({
        "por_dia": [(f, o, n, g, m, sc) for f, o, n, g, m, sc in resumen_dia],
        "agg":     {k: dict(v) for k, v in agg.items()},
        "totales": {"ords": total_ords, "neto": total_neto, "gan": total_gan, "margen": margen_p},
    }, ensure_ascii=False, indent=2),
    encoding="utf-8",
)
print("\nDatos guardados en data/semanal_data.json")
