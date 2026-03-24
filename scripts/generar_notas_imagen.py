"""
Genera imágenes PNG tipo tarjeta de notas de pedido por proveedor.
Lee el General XLSX corregido y produce un PNG por proveedor para WhatsApp.

Uso:
  python scripts/generar_notas_imagen.py data/xlsx/Nota_Pedido_General_2026-03-19.xlsx 2026-03-19
"""

import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from PIL import Image, ImageDraw, ImageFont

# --- Colores por proveedor ---
COLORES = {
    "Chapetes": {"header": "#2E7D6F", "badge_bg": "#D4EDDA", "badge_text": "#2E7D6F"},
    "Dartacan": {"header": "#1B3A5C", "badge_bg": "#D6E4F0", "badge_text": "#1B3A5C"},
    "Invet":    {"header": "#5B3A8C", "badge_bg": "#E8D5F5", "badge_text": "#5B3A8C"},
    "Costco":   {"header": "#D4651E", "badge_bg": "#FDEBD0", "badge_text": "#D4651E"},
}

DEFAULT_COLOR = {"header": "#444444", "badge_bg": "#E0E0E0", "badge_text": "#444444"}

# --- Dimensiones ---
WIDTH = 800
HEADER_H = 200
ROW_H = 80
FOOTER_H = 80
PADDING_X = 40
BADGE_SIZE = 48

# --- Meses en español ---
MESES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def hex_to_rgb(hex_color: str) -> tuple:
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def format_fecha_espanol(fecha_iso: str) -> str:
    """'2026-03-19' → '19 de marzo, 2026'"""
    dt = datetime.strptime(fecha_iso, "%Y-%m-%d")
    return f"{dt.day} de {MESES[dt.month]}, {dt.year}"


def get_fonts():
    """Carga Segoe UI (Windows system font) con fallback."""
    font_paths = [
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    bold_paths = [
        "C:/Windows/Fonts/segoeuib.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    font_regular = None
    font_bold = None
    for p in font_paths:
        if Path(p).exists():
            font_regular = p
            break
    for p in bold_paths:
        if Path(p).exists():
            font_bold = p
            break

    return {
        "header_big": ImageFont.truetype(font_bold or font_regular, 36),
        "header_proveedor": ImageFont.truetype(font_bold or font_regular, 48),
        "header_fecha": ImageFont.truetype(font_regular, 24),
        "col_header": ImageFont.truetype(font_bold or font_regular, 22),
        "row_text": ImageFont.truetype(font_regular, 24),
        "badge_text": ImageFont.truetype(font_bold or font_regular, 22),
        "footer_text": ImageFont.truetype(font_bold or font_regular, 24),
    }


def leer_general_xlsx(path: str) -> dict:
    """Lee el General XLSX y agrupa por proveedor.
    Cada hoja = un proveedor. Fila 1=titulo, Fila 2=headers, Datos desde fila 3.
    Detecta columnas por nombre para no depender del orden.

    Returns: {proveedor: [{"producto": str, "cantidad": int}, ...]}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    resultado = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        proveedor = sheet_name.strip()
        productos = []

        # Leer fila 2 para detectar posición de columnas por nombre
        rows_all = list(ws.iter_rows(values_only=True))
        if len(rows_all) < 2:
            continue

        headers_row = [str(h).strip().lower() if h else "" for h in rows_all[1]]
        # Buscar índice de "producto" y "cantidad"
        idx_prod = next(
            (i for i, h in enumerate(headers_row) if "producto" in h or "product" in h), None
        )
        idx_cant = next(
            (i for i, h in enumerate(headers_row) if "cantidad" in h or "qty" in h or "cant" in h), None
        )

        # Fallback: si no hay headers claros, asumir A=Producto, B=Cantidad
        if idx_prod is None:
            idx_prod = 0
        if idx_cant is None:
            idx_cant = 1

        # Datos desde fila 3 (índice 2)
        for row in rows_all[2:]:
            if not row or len(row) <= max(idx_prod, idx_cant):
                continue
            producto = row[idx_prod]
            cantidad = row[idx_cant]
            if not producto or str(producto).strip().lower() in ("producto", ""):
                continue
            try:
                cantidad = int(cantidad)
            except (ValueError, TypeError):
                cantidad = 1
            productos.append({"producto": str(producto).strip(), "cantidad": cantidad})

        if productos:
            resultado[proveedor] = productos

    wb.close()
    return resultado


def render_nota_imagen(proveedor: str, productos: list, fecha: str, output_path: str):
    """Genera un PNG con la nota de pedido para un proveedor."""
    colores = COLORES.get(proveedor, DEFAULT_COLOR)
    header_rgb = hex_to_rgb(colores["header"])
    badge_bg_rgb = hex_to_rgb(colores["badge_bg"])
    badge_text_rgb = hex_to_rgb(colores["badge_text"])
    fecha_formateada = format_fecha_espanol(fecha)
    fonts = get_fonts()

    n_productos = len(productos)
    total_h = HEADER_H + 50 + (ROW_H * n_productos) + FOOTER_H
    total_piezas = sum(p["cantidad"] for p in productos)

    img = Image.new("RGB", (WIDTH, total_h), "white")
    draw = ImageDraw.Draw(img)

    # --- Header ---
    draw.rectangle([(0, 0), (WIDTH, HEADER_H)], fill=header_rgb)
    draw.text((PADDING_X, 30), "PEDIDO A PROVEEDOR", fill="white", font=fonts["header_big"])
    draw.text((PADDING_X, 75), proveedor, fill="white", font=fonts["header_proveedor"])
    draw.text((PADDING_X, 140), fecha_formateada, fill=(200, 220, 230), font=fonts["header_fecha"])

    # --- Column headers ---
    y = HEADER_H + 15
    draw.text((PADDING_X, y), "Producto", fill="#666666", font=fonts["col_header"])
    draw.text((WIDTH - PADDING_X - 60, y), "Qty", fill="#666666", font=fonts["col_header"])
    y += 35

    # --- Rows ---
    for i, prod in enumerate(productos):
        row_y = y + (ROW_H * i)
        # Separator line
        draw.line([(PADDING_X, row_y), (WIDTH - PADDING_X, row_y)], fill="#E0E0E0", width=1)
        # Product name
        text_y = row_y + (ROW_H - 24) // 2
        draw.text((PADDING_X, text_y), prod["producto"], fill="#333333", font=fonts["row_text"])
        # Badge
        badge_x = WIDTH - PADDING_X - BADGE_SIZE
        badge_y = row_y + (ROW_H - BADGE_SIZE) // 2
        draw.rounded_rectangle(
            [(badge_x, badge_y), (badge_x + BADGE_SIZE, badge_y + BADGE_SIZE)],
            radius=10,
            fill=badge_bg_rgb,
        )
        qty_str = str(prod["cantidad"])
        bbox = draw.textbbox((0, 0), qty_str, font=fonts["badge_text"])
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text(
            (badge_x + (BADGE_SIZE - tw) // 2, badge_y + (BADGE_SIZE - th) // 2 - 2),
            qty_str,
            fill=badge_text_rgb,
            font=fonts["badge_text"],
        )

    # --- Footer ---
    footer_y = y + (ROW_H * n_productos)
    draw.line([(PADDING_X, footer_y), (WIDTH - PADDING_X, footer_y)], fill="#CCCCCC", width=2)
    footer_text_y = footer_y + (FOOTER_H - 24) // 2
    draw.text((PADDING_X, footer_text_y), "Total piezas", fill="#333333", font=fonts["footer_text"])
    total_str = str(total_piezas)
    bbox = draw.textbbox((0, 0), total_str, font=fonts["footer_text"])
    tw = bbox[2] - bbox[0]
    draw.text(
        (WIDTH - PADDING_X - tw, footer_text_y),
        total_str,
        fill=header_rgb,
        font=fonts["footer_text"],
    )

    img.save(output_path, "PNG")
    print(f"  Generado: {output_path}")


def main():
    if len(sys.argv) < 3:
        print("Uso: python scripts/generar_notas_imagen.py <General_XLSX> <YYYY-MM-DD>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    fecha = sys.argv[2]

    if not Path(xlsx_path).exists():
        print(f"ERROR: No existe {xlsx_path}")
        sys.exit(1)

    print(f"Leyendo {xlsx_path}...")
    datos = leer_general_xlsx(xlsx_path)

    if not datos:
        print("ERROR: No se encontraron datos en el XLSX.")
        sys.exit(1)

    output_dir = Path(xlsx_path).parent
    print(f"Generando PNGs para {len(datos)} proveedor(es)...\n")

    for proveedor, productos in datos.items():
        safe_name = proveedor.replace(" ", "_")
        output_path = str(output_dir / f"Nota_Pedido_{safe_name}_{fecha}.png")
        render_nota_imagen(proveedor, productos, fecha, output_path)

    print(f"\nListo. {len(datos)} imagen(es) generadas en {output_dir}")


if __name__ == "__main__":
    main()
