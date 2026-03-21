"""Genera Reporte_Semanal_S12_2026.xlsx — Semana 16-20 Mar 2026.

Fuentes:
- 03-17/18: costos embebidos en hoja Ventas de cada reporte
- 03-19: costos embebidos + Sheet para los 9 SIN COSTO (ya registrados)
- 03-20: costos cruzados desde Sheet (el reporte solo tiene Ventas sin costo)
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

from scripts.catalogo import leer_catalogo

# ── PALETA ─────────────────────────────────────────────────────────────────────
H_FILL = PatternFill("solid", fgColor="2C3E50")
S_FILL = PatternFill("solid", fgColor="34495E")
V_FILL = PatternFill("solid", fgColor="27AE60")
A_FILL = PatternFill("solid", fgColor="F39C12")
R_FILL = PatternFill("solid", fgColor="E74C3C")
G_FILL = PatternFill("solid", fgColor="ECF0F1")
H_FONT = Font(bold=True, color="FFFFFF")
B_FONT = Font(bold=True)
CTR    = Alignment(horizontal="center")

DART = "Dartacan"; CHAP = "Chapetes"; INVE = "Invet"; COST = "Costco"


def sem_fill(m):
    if m is None: return PatternFill("solid", fgColor="BDC3C7")
    return V_FILL if m >= 14 else (A_FILL if m >= 8 else R_FILL)


def sem_txt(m):
    if m is None: return "SIN COSTO"
    return f"VERDE {m:.1f}%" if m >= 14 else (f"AMARI {m:.1f}%" if m >= 8 else f"ROJO  {m:.1f}%")


def hrow(ws, r, vals, fill=H_FILL, font=H_FONT):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.fill = fill; cell.font = font; cell.alignment = CTR


# ── CATÁLOGO SHEET ─────────────────────────────────────────────────────────────
cat = leer_catalogo()
SHEET = {v["producto"].lower(): float(v["costo_actual"]) for v in cat.values()
         if v.get("costo_actual")}


def costo_exacto(nombre):
    return SHEET.get(nombre.lower())


# ── DATOS DIARIOS (Resumen de cada reporte = fuente oficial) ───────────────────
# (fecha, ordenes, neto, ganancia, margen, notas)
POR_DIA_RESUMEN = [
    ("2026-03-17",  5, 4178.55,  793.55, 19.0, "5 ords — costos Dartacan #363 registrados"),
    ("2026-03-18", 18, 14528.73, 1345.91, 9.9, "18 ords — Silver Kan SIN COSTO en reporte"),
    ("2026-03-19", 14, 8845.96,  973.94, 11.0, "14 ords — 9 SIN COSTO corregidos con Sheet"),
    ("2026-03-20", 18, 14920.06, 1691.21, 11.3, "18 ords — Kan Kan $380 corregido"),
]
# 03-19 corregido: $240.44 original + $731.30 de los 9 SIN COSTO ahora costeados
# Silver Kan x4 ($491+$638+$352split+$631) = $312 gan; Chapetes varios = $390+$29.30

# ── PRODUCTOS (ventas detalle — costos del reporte o Sheet) ────────────────────
# (prod, u, neto_total, costo_total, proveedor)
VENTAS = [
    # ─── 03-17 ────────────────────────────────────────────────────────────────
    ("Perron Adulto 25kg",              1,  788.30,   535.00, DART),
    ("Pedigree 20kg Res/Veg",           1,  946.98,   725.00, DART),
    ("Chapetes Premium 18kg",           1,  501.19,   410.00, CHAP),
    ("Ganador Premium 20kg",            1, 1164.10,   990.00, DART),
    ("Pedigree 20kg Res/Veg",           1,  777.98,   725.00, DART),
    # ─── 03-18 ────────────────────────────────────────────────────────────────
    ("Silver Kan 25kg",                 2,  982.60,   900.00, DART),   # $450x2, SIN COSTO en rep.
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("Maskottchen Premium 15kg",        1,  717.00,   525.00, CHAP),
    ("Maskottchen Premium 15kg",        1,  717.00,   525.00, CHAP),
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("Chapetes 20kg Amarillos",         1,  531.45,   300.00, CHAP),
    ("Chapetes 20kg Amarillos",         1,  333.25,   300.00, CHAP),
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("LiveClear Gato 3.18kg",           1,  798.00,   768.61, INVE),
    ("Arena Scoop Away 19kg 2-Pack",    1,  895.97,   798.00, COST),
    ("Pedigree 20kg Res/Veg",           1,  778.00,   725.00, DART),
    ("Dog Chow Adulto 25kg",            1,  954.75,   900.00, DART),
    ("Pedigree 20kg Res/Veg",           1,  778.00,   725.00, DART),
    ("Gatina 15kg",                     1,  523.67,   495.00, DART),
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("LiveClear Gato 3.18kg",           1,  798.00,   768.61, INVE),
    ("6 Latas ProPlan Gastro 380g",     1,  454.49,   418.55, INVE),
    # ─── 03-19 ────────────────────────────────────────────────────────────────
    ("Silver Kan 25kg",                 1,  491.00,   450.00, DART),   # Flex regular
    ("Silver Kan 25kg",                 1,  631.00,   450.00, DART),   # posible Flex
    ("Silver Kan 25kg",                 2,  990.00,   900.00, DART),   # split $638+$352
    ("Chapetes Super Premium Gato 5kg", 1,  338.00,   185.00, CHAP),
    ("2 Costales Chapetes Gato 15kg",   2,  978.00,   880.00, CHAP),   # 2x$440
    ("24 Sobres Minino Plus",          24,  224.00,   204.00, DART),   # 24x$8.50
    ("Chapetes Premium 18kg",           1,  501.19,   410.00, CHAP),
    ("Perron Adulto 25kg",              1,  558.00,   535.00, DART),
    ("6 Latas ProPlan Gastro 380g",     1,  454.00,   418.55, INVE),
    ("3 Costales Chapetes Premium 18kg",3, 1349.00,  1230.00, CHAP),   # 3x$410
    ("Gatina 15kg",                     1,  523.67,   495.00, DART),
    ("Ganador Premium 20kg",            1, 1053.60,   990.00, DART),
    ("Campeon Adulto 25kg",             1,  754.30,   725.00, DART),
    # ─── 03-20 ────────────────────────────────────────────────────────────────
    ("Kan Kan 25kg",                    1,  622.51,   380.00, DART),
    ("Campeon Adulto 25kg",             1,  754.30,   725.00, DART),
    ("6 Latas ProPlan Gastro 380g",     1,  478.95,   418.55, INVE),
    ("Campeon Adulto 25kg",             1,  754.30,   725.00, DART),
    ("Chapetes Premium 18kg",           1,  501.19,   410.00, CHAP),
    ("Dog Chow Adulto 25kg",            1,  954.75,   900.00, DART),
    ("Kirkland Salmon/Camote 15.87kg",  1, 1064.62,   781.15, COST),
    ("Perron X2 Costales 50kg",         1, 1149.76,  1070.00, DART),
    ("Perron X2 Costales 50kg",         1, 1149.76,  1070.00, DART),
    ("Dog Chow Adulto 25kg",            1,  954.75,   900.00, DART),
    ("Campeon Adulto 25kg",             1,  754.30,   725.00, DART),
    ("Perron X2 Costales 50kg",         1, 1149.76,  1070.00, DART),
    ("Perron X2 Costales 50kg",         1, 1149.76,  1070.00, DART),
    ("Campeon Adulto 25kg",             1,  754.30,   725.00, DART),
    ("Gatina 15kg",                     1,  523.67,   495.00, DART),
    ("Arena Scoop Away 19kg 2-Pack",    1,  895.97,   798.00, COST),
    ("Chapetes Super Premium Gato 5kg", 1,  242.79,   185.00, CHAP),
    ("Kirkland Salmon/Camote 15.87kg",  1, 1064.62,   781.15, COST),
]

# Consolidar por producto
agg = defaultdict(lambda: {"u": 0, "neto": 0.0, "costo": 0.0, "prov": ""})
for prod, u, neto, costo, prov in VENTAS:
    agg[prod]["u"]     += u
    agg[prod]["neto"]  += neto
    agg[prod]["costo"] += costo
    agg[prod]["prov"]   = prov

total_ords = sum(d[1] for d in POR_DIA_RESUMEN)
total_neto = sum(d[2] for d in POR_DIA_RESUMEN)
total_gan  = sum(d[3] for d in POR_DIA_RESUMEN)
margen_p   = total_gan / total_neto * 100

# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── HOJA 1: RESUMEN ───────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Resumen"
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 22

ws1.merge_cells("A1:D1")
ws1["A1"] = "CROQUETERIA GABY  |  Reporte Semanal S12 2026"
ws1["A1"].font = Font(bold=True, size=14, color="2C3E50")
ws1.merge_cells("A2:D2")
ws1["A2"] = "Semana 16 - 20 Marzo 2026  (4 dias con ventas)"
ws1["A2"].font = Font(italic=True, color="7F8C8D")

metricas = [
    ("Ordenes totales",      str(total_ords),             None),
    ("Neto total (MXN)",     f"${total_neto:,.2f}",       "27AE60"),
    ("Ganancia total",       f"${total_gan:,.2f}",        "27AE60"),
    ("Margen promedio pond.", f"{margen_p:.1f}%",          "2980B9"),
]
for i, (lbl, val, color) in enumerate(metricas, 4):
    ws1.cell(i, 1, lbl).font = B_FONT
    cell = ws1.cell(i, 2, val)
    if color:
        cell.font = Font(bold=True, size=12, color=color)

# ── HOJA 2: POR DÍA ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Por_Dia")
for col, w in zip("ABCDEF", [14, 10, 16, 16, 12, 50]):
    ws2.column_dimensions[col].width = w

hrow(ws2, 1, ["Fecha", "Ordenes", "Neto MXN", "Ganancia", "Margen", "Notas"])
for i, (fecha, ords, neto, gan, marg, nota) in enumerate(POR_DIA_RESUMEN, 2):
    ws2.cell(i, 1, fecha)
    ws2.cell(i, 2, ords).alignment = CTR
    ws2.cell(i, 3, f"${neto:,.2f}")
    ws2.cell(i, 4, f"${gan:,.2f}")
    c = ws2.cell(i, 5, f"{marg:.1f}%")
    c.fill = sem_fill(marg); c.font = Font(bold=True, color="FFFFFF"); c.alignment = CTR
    ws2.cell(i, 6, nota).font = Font(italic=True, color="7F8C8D")

r = len(POR_DIA_RESUMEN) + 2
hrow(ws2, r, ["TOTAL", str(total_ords), f"${total_neto:,.2f}",
              f"${total_gan:,.2f}", f"{margen_p:.1f}%", ""], fill=S_FILL)

# ── HOJA 3: TOP PRODUCTOS ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("Top_Productos")
for col, w in zip("ABCDEFG", [44, 8, 16, 16, 14, 12, 18]):
    ws3.column_dimensions[col].width = w

hrow(ws3, 1, ["Producto", "Units", "Neto Total", "Costo Total", "Ganancia", "Proveedor", "Semaforo"])
for i, (prod, d) in enumerate(
        sorted(agg.items(), key=lambda x: x[1]["neto"], reverse=True), 2):
    n, c = d["neto"], d["costo"]
    gan = n - c
    m   = gan / n * 100
    fill = G_FILL if i % 2 == 0 else PatternFill()
    ws3.cell(i, 1, prod).fill = fill
    ws3.cell(i, 2, d["u"]).alignment = CTR
    ws3.cell(i, 3, f"${n:,.2f}").fill = fill
    ws3.cell(i, 4, f"${c:,.2f}").fill = fill
    ws3.cell(i, 5, f"${gan:,.2f}").fill = fill
    ws3.cell(i, 6, d["prov"]).fill = fill
    cell = ws3.cell(i, 7, sem_txt(m))
    cell.fill = sem_fill(m); cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = CTR

# ── HOJA 4: ALERTAS ───────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Alertas")
for col, w in zip("ABCDEFG", [44, 8, 12, 12, 12, 12, 42]):
    ws4.column_dimensions[col].width = w

hrow(ws4, 1, ["Producto", "Units", "Neto/u", "Costo/u", "Margen", "Proveedor", "Accion sugerida"])

ACCIONES = {
    "Campeon Adulto 25kg":         "Repreciar ~$1,000 lista — 5 vendidos, margen 3.9%",
    "LiveClear Gato 3.18kg":       "Repreciar ~$1,090 lista — margen 3.6%",
    "Gatina 15kg":                 "Repreciar ~$700 lista — 3 vendidos",
    "Perron Adulto 25kg":          "Repreciar ~$735 lista — margen 4.2%",
    "Dog Chow Adulto 25kg":        "Repreciar ~$1,230 lista — 3 vendidos",
    "Pedigree 20kg Res/Veg":       "Repreciar ~$980 lista — margen 6.8%",
    "Ganador Premium 20kg":        "Repreciar ~$1,350 lista — 7 vendidos (el mas vendido!)",
    "Perron X2 Costales 50kg":     "Revisar precio paquete x2 — 4 paquetes vendidos",
    "Chapetes 20kg Amarillos":     "Margen bajo en 1 de 2 ventas (envio diferente)",
    "6 Latas ProPlan Gastro 380g": "Monitorear — margen limite 7.5%",
}

rojos = [(p, d) for p, d in agg.items()
         if d["costo"] > 0 and (d["neto"] - d["costo"]) / d["neto"] * 100 < 8]
rojos.sort(key=lambda x: (x[1]["neto"] - x[1]["costo"]) / x[1]["neto"])

for i, (prod, d) in enumerate(rojos, 2):
    nu = d["neto"] / d["u"]
    cu = d["costo"] / d["u"]
    m  = (nu - cu) / nu * 100
    ws4.cell(i, 1, prod).font = B_FONT
    ws4.cell(i, 2, d["u"]).alignment = CTR
    ws4.cell(i, 3, f"${nu:,.2f}")
    ws4.cell(i, 4, f"${cu:,.2f}")
    c5 = ws4.cell(i, 5, f"{m:.1f}%")
    c5.fill = R_FILL; c5.font = Font(bold=True, color="FFFFFF"); c5.alignment = CTR
    ws4.cell(i, 6, d["prov"])
    ws4.cell(i, 7, ACCIONES.get(prod, "Revisar precio")).font = Font(italic=True, color="C0392B")

# ── HOJA 5: COMPRAS SUGERIDAS ─────────────────────────────────────────────────
ws5 = wb.create_sheet("Compras_Sugeridas")
for col, w in zip("ABCDE", [44, 12, 14, 12, 44]):
    ws5.column_dimensions[col].width = w

hrow(ws5, 1, ["Producto", "Vendidos/sem", "Costo unit", "Proveedor", "Nota"])

NOTAS_COMPRA = {
    "Ganador Premium 20kg":         "7 vendidos — REPRECIAR antes de restock urgente",
    "Campeon Adulto 25kg":          "5 vendidos — margen 3.9% — REPRECIAR PRIMERO",
    "Perron X2 Costales 50kg":      "4 paquetes (8 costales) — revisar margen antes",
    "Chapetes Premium 18kg":        "3 vendidos — margen 18.2% — mantener stock",
    "Pedigree 20kg Res/Veg":        "3 vendidos — repreciar antes de restock",
    "Gatina 15kg":                  "3 vendidos — repreciar antes de restock",
    "Dog Chow Adulto 25kg":         "3 vendidos — repreciar antes de restock",
    "Silver Kan 25kg":              "4 vendidos — costo provisional $450",
    "Arena Scoop Away 19kg 2-Pack": "2 vendidos — margen 10.9% ok",
    "Kirkland Salmon/Camote 15.87kg": "2 vendidos Flex — margen 26.6% excelente",
    "LiveClear Gato 3.18kg":        "2 vendidos — margen 3.6% — REPRECIAR PRIMERO",
    "Maskottchen Premium 15kg":     "2 vendidos — margen 26.8% excelente",
}

restock = sorted(
    [(p, d) for p, d in agg.items()],
    key=lambda x: x[1]["u"], reverse=True
)
for i, (prod, d) in enumerate(restock, 2):
    cu = d["costo"] / d["u"] if d["u"] else 0
    ws5.cell(i, 1, prod)
    ws5.cell(i, 2, d["u"]).alignment = CTR
    ws5.cell(i, 3, f"${cu:,.2f}" if cu else "—")
    ws5.cell(i, 4, d["prov"])
    nota = NOTAS_COMPRA.get(prod, "Monitorear")
    c = ws5.cell(i, 5, nota)
    if "REPRECIAR" in nota:
        c.font = Font(bold=True, italic=True, color="C0392B")
    else:
        c.font = Font(italic=True, color="2C3E50")

# ── GUARDAR ───────────────────────────────────────────────────────────────────
out = Path("data/xlsx/Reporte_Semanal_S12_2026.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out))

print(f"Generado: {out}")
print(f"  {total_ords} ordenes | ${total_neto:,.2f} neto | ${total_gan:,.2f} ganancia | {margen_p:.1f}% margen")
print(f"  Dias: 17-20 Mar 2026 | {len(agg)} productos unicos")
